from app import app, mysql, is_logged_in
from flask import render_template, request, redirect, url_for, session, flash, send_file
from models.datasets import Dataset
from werkzeug.security import generate_password_hash, check_password_hash
from pandas import DataFrame
from datetime import datetime
from openpyxl import Workbook

@app.route('/datasets')
@is_logged_in()
def index_dataset():
    # list datasets        
    datasets = Dataset.get_all(mysql)

    # return datasets with index.html
    return render_template('datasets/index.html', datasets=datasets)

@app.route('/datasets/upload', methods=['GET'])
def upload_form():
    # return upload.html
    return render_template('datasets/upload.html')

@app.route('/datasets/upload', methods=['POST'])
def upload_dataset():
    # get file from form
    file = request.files['file']

    # save file to database
    Dataset.store_csv(mysql, file)

    # redirect to index
    return redirect(url_for('index_dataset'))

@app.route('/datasets/delete_all')
def delete_all_dataset():
    # delete all datasets
    Dataset.delete_all(mysql)

    # return datasets with index.html
    return redirect(url_for('index_dataset'))

@app.route('/datasets/<int:id>/edit')
@is_logged_in()
def edit_dataset(id):
    # get dataset by id
    dataset = Dataset.get_by_id(mysql, id)
    
    # return edit form with dataset
    return render_template('datasets/edit.html', dataset=dataset)

@app.route('/datasets/<int:id>/update', methods=['POST'])
@is_logged_in()
def update_dataset(id):
    # get form data
    code = request.form['code']
    date = request.form['date']
    items = request.form['items']
    total = request.form['total']
    
    # update dataset
    Dataset.update(mysql, id, code, date, items, total)
    
    # redirect to index
    flash('Dataset berhasil diupdate', 'success')
    return redirect(url_for('index_dataset'))

@app.route('/datasets/export')
def export_dataset():
    # get all datasets
    datasets = Dataset.get_all(mysql)

    # group datasets by code and get unique items
    grouped_datasets = {}
    grouped_date = {}
    for dataset in datasets:
        if dataset['code'] not in grouped_datasets:            
            grouped_datasets[dataset['code']] = dataset['items'].split(',')
            grouped_date[dataset['code']] = dataset['date']
        else:
            grouped_datasets[dataset['code']] += dataset['items'].split(',')
            grouped_date[dataset['code']] = dataset['date']

    # count items for each code and each item
    for code in grouped_datasets:
        grouped_datasets[code] = dict((i, grouped_datasets[code].count(i)) for i in grouped_datasets[code])
            
    # sort items by name
    for code in grouped_datasets:
        grouped_datasets[code] = dict(sorted(grouped_datasets[code].items()))    

    # get all unique items
    unique_items = []
    for code in grouped_datasets:
        unique_items += list(grouped_datasets[code].keys())

    # get unique items
    unique_items = list(set(unique_items))

    # sort unique items by name
    unique_items.sort()    

    # create excel file
    wb = Workbook()

    # set sheet
    sheet = wb.active
    sheet.title = 'Data Penjualan Barang'

    # set header
    sheet['A1'] = 'Kode'
    sheet['B1'] = 'Tanggal'

    for i in range(len(unique_items)):
        sheet.cell(row=1, column=i+3).value = unique_items[i]

    # set data
    row = 2
    for code in grouped_datasets:
        sheet.cell(row=row, column=1).value = code
        sheet.cell(row=row, column=2).value = grouped_date[code]
        for i in range(len(unique_items)):
            if unique_items[i] in grouped_datasets[code]:
                sheet.cell(row=row, column=i+3).value = grouped_datasets[code][unique_items[i]]
            else:
                sheet.cell(row=row, column=i+3).value = 0
        row += 1

    # File name
    file_name = 'Data Penjualan Barang ' + datetime.now().strftime('%d-%m-%Y %H-%M-%S') + '.xlsx'

    # save file
    wb.save(file_name) 

    # return download file
    return send_file(file_name, as_attachment=True)